from django.shortcuts import render, redirect
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from Solicitudes.forms import SolicitudForm, SolicitudFilterForm, RespuestaForm
from Solicitudes.models import Solicitud
from django.shortcuts import render, get_object_or_404
from django.contrib import messages
from django.utils.dateparse import parse_date
from .serializers import SolicitudSerializer
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.http import Http404
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.tokens import RefreshToken
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


@login_required(login_url='index')
def crear_solicitud(request):
    form = SolicitudForm(request.POST)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            return redirect('../../Solicitudes/Solicitud_Creada')

    return render(request, 'Solicitudes/crearSolicitud.html', {'form': form})


@login_required(login_url='index')
def responder_solicitud(request):
    form = SolicitudFilterForm(request.GET)
    solicitudes = Solicitud.objects.all()

    if form.is_valid():
        nombre = form.cleaned_data.get('nombre')
        estado = form.cleaned_data.get('estado')
        municipio = form.cleaned_data.get('municipio')
        tipo_problema = form.cleaned_data.get('tipo_problema')
        resuelto = form.cleaned_data.get('resuelto')
        fecha = form.cleaned_data.get('fecha')

        if nombre:
            solicitudes = solicitudes.filter(nombre__icontains=nombre)
        if estado:
            solicitudes = solicitudes.filter(estado__icontains=estado)
        if municipio:
            solicitudes = solicitudes.filter(municipio__icontains=municipio)
        if resuelto is not None and resuelto != "":
            solicitudes = solicitudes.filter(resuelto=resuelto)
        if tipo_problema:
            solicitudes = solicitudes.filter(
                tipo_problema__icontains=tipo_problema)
        if fecha:
            fecha_parsed = parse_date(str(fecha))
            if fecha_parsed:
                solicitudes = solicitudes.filter(
                    fecha_creacion__date=fecha_parsed)
        if not solicitudes.exists():
            messages.info(
                request, 'No se encontraron resultados con los filtros seleccionados.')

    if request.method == 'POST':
        action = request.POST.get('action')

        if action:
            if action.startswith('eliminar_'):
                solicitud_id = action.replace('eliminar_', '')
                solicitud = get_object_or_404(Solicitud, id=solicitud_id)
                solicitud.delete()
                messages.success(request, 'La solicitud ha sido eliminada.')

                return redirect('Solicitudes:responder_solicitud')

    return render(request, 'Solicitudes/responderSolicitud.html', {'solicitudes': solicitudes, 'form': form})


@login_required(login_url='index')
def lista_solicitud(request):
    form = SolicitudFilterForm(request.GET)
    solicitudes = Solicitud.objects.all()

    if form.is_valid():
        nombre = form.cleaned_data.get('nombre')
        estado = form.cleaned_data.get('estado')
        municipio = form.cleaned_data.get('municipio')
        tipo_problema = form.cleaned_data.get('tipo_problema')
        resuelto = form.cleaned_data.get('resuelto')
        fecha = form.cleaned_data.get('fecha')

        if nombre:
            solicitudes = solicitudes.filter(nombre__icontains=nombre)
        if estado:
            solicitudes = solicitudes.filter(estado__icontains=estado)
        if municipio:
            solicitudes = solicitudes.filter(municipio__icontains=municipio)
        if resuelto is not None and resuelto != "":
            solicitudes = solicitudes.filter(resuelto=resuelto)
        if tipo_problema:
            solicitudes = solicitudes.filter(
                tipo_problema__icontains=tipo_problema)
        if fecha:
            fecha_parsed = parse_date(str(fecha))
            if fecha_parsed:
                solicitudes = solicitudes.filter(
                    fecha_creacion__date=fecha_parsed)
        if not solicitudes.exists():
            messages.info(
                request, 'No se encontraron resultados con los filtros seleccionados.')

    return render(request, 'Solicitudes/listaSolicitudes.html', {'solicitudes': solicitudes, 'form': form})


@login_required(login_url='index')
def solicitud_Creada(request):
    return render(request, 'Solicitudes/solicitudCreada.html', {})


@login_required(login_url='index')
def detalles_solicitud(request, solicitud_id):
    solicitud = get_object_or_404(Solicitud, id=solicitud_id)
    return render(request, 'Solicitudes/detalles_solicitud.html', {'solicitud': solicitud})


@login_required(login_url='index')
def editar_solicitud(request, solicitud_id):
    solicitud = get_object_or_404(Solicitud, id=solicitud_id)

    if request.method == 'POST':
        form = SolicitudForm(request.POST, instance=solicitud)
        if form.is_valid():
            form.save()
            messages.success(request, 'La solicitud ha sido actualizada.')
            return redirect('Solicitudes:lista_solicitudes')
    else:
        form = SolicitudForm(instance=solicitud)

    return render(request, 'Solicitudes/editar_Solicitud.html', {'form': form, 'solicitud': solicitud})


@login_required(login_url='index')
def respuesta_de_solicitud(request, solicitud_id):
    solicitud = get_object_or_404(Solicitud, id=solicitud_id)

    if request.method == 'POST':
        respuesta_form = RespuestaForm(
            request.POST, request.FILES, instance=solicitud)

        if respuesta_form.is_valid():
            respuesta_form.save()

            # Verificar si los campos de respuesta han sido completados
            if (respuesta_form.cleaned_data['nombre_respuesta'] and
                    respuesta_form.cleaned_data['puesto_respuesta'] and
                    respuesta_form.cleaned_data['detalle_respuesta']):
                solicitud.resuelto = True
                solicitud.save()
            messages.success(request, 'La solicitud ha sido respondida.')
            return redirect('Solicitudes:responder_solicitud')

        else:
            print(respuesta_form.errors)
    else:
        respuesta_form = RespuestaForm(instance=solicitud)

    return render(request, 'Solicitudes/respuesta_de_solicitud.html', {'respuesta_form': respuesta_form, 'solicitud': solicitud})


class SolicitudListaView(APIView):
    def get(self, request):
        solicitudes = Solicitud.objects.all()
        serializer = SolicitudSerializer(solicitudes, many=True)
        return Response(serializer.data)

    def post(self, request):
        serializer = SolicitudSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class SolicitudDetalleView(APIView):
    permission_classes = [IsAuthenticated]

    def get_object(self, pk):
        try:
            return Solicitud.objects.get(pk=pk)
        except Solicitud.DoesNotExist:
            raise Http404

    def get(self, request, pk):
        solicitud = self.get_object(pk)
        serializer = SolicitudSerializer(solicitud)
        return Response(serializer.data)

    def put(self, request, pk):
        solicitud = self.get_object(pk)
        serializer = SolicitudSerializer(solicitud, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def patch(self, request, pk):
        solicitud = self.get_object(pk)
        serializer = SolicitudSerializer(
            solicitud, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk):
        solicitud = self.get_object(pk)
        solicitud.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class TokenObtainView(APIView):
    def post(self, request):
        username = request.data.get('username')
        password = request.data.get('password')

        if username is None or password is None:
            return Response({'error': 'Se requiere nombre de usuario y contraseña.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            user = User.objects.get(username=username)
            if user.check_password(password):
                refresh = RefreshToken.for_user(user)
                return Response({
                    'access_token': str(refresh.access_token),
                    'refresh_token': str(refresh),
                })
        except User.DoesNotExist:
            pass

        return Response({'error': 'Credenciales inválidas.'}, status=status.HTTP_401_UNAUTHORIZED)


def generar_reporte_excel(request):
    solicitudes = Solicitud.objects.all()

    wb = Workbook()
    ws = wb.active

    header_style = Alignment(
        horizontal='center', vertical='center', wrap_text=True)
    header_font = Font(bold=True)

    headers = ["Fecha de Creación", "Resuelto", "Nombre", "Tipo de Problema", "Tipo Otro",
               "Estado", "Municipio", "Dirección", "Número Exterior", "Número Interior",
               "Código Postal", "Teléfono", "Correo Electrónico", "Descripción",
               "Nombre de Respuesta", "Puesto de Respuesta", "Detalle de Respuesta"]

    for col_num, header_text in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        header_cell = ws[f"{col_letter}1"]
        header_cell.value = header_text
        header_cell.alignment = header_style
        header_cell.font = header_font

    cell_style = Alignment(horizontal='center',
                           vertical='center', wrap_text=True)

    for solicitud in solicitudes:
        fecha_creacion_no_tz = solicitud.fecha_creacion.replace(tzinfo=None)

        row = [
            fecha_creacion_no_tz, solicitud.resuelto, solicitud.nombre,
            solicitud.tipo_problema, solicitud.tipo_otro, solicitud.estado,
            solicitud.municipio, solicitud.direccion, solicitud.num_ext,
            solicitud.num_int, solicitud.codigo_postal, solicitud.telefono,
            solicitud.email, solicitud.descripcion, solicitud.nombre_respuesta,
            solicitud.puesto_respuesta, solicitud.detalle_respuesta
        ]

        ws.append(row)
        for col_num in range(1, len(row) + 1):
            col_letter = get_column_letter(col_num)
            cell = ws[f"{col_letter}{ws.max_row}"]
            cell.alignment = cell_style

    for col_num, _ in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].auto_size = True

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reporte_solicitudes.xlsx'

    wb.save(response)

    return response
