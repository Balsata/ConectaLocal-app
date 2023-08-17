from django.shortcuts import render
from open311.models import ListaOpen311
import requests
import json
from django.http import JsonResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from open311.forms import SolicitudFilterForm
from django.utils.dateparse import parse_date
from django.core.paginator import Paginator
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def solicitud(request):
    context = {}
    for page in range(11, 31):

        data = {
            "lat": 41.3083,
            "long": -72.9279,
            "page_size": 100,
            "page": page
        }
        r = requests.get(
            'https://seeclickfix.com/open311/v2/requests.json/', data=data)
        response = json.loads(r.content)

        for resp in response:
            try:
                if not ListaOpen311.objects.filter(request_id=resp['service_request_id']).exists():
                    solicitud = ListaOpen311(descripcion=resp['description'], request_id=resp['service_request_id'], address=resp['address'],
                                             lat=resp['lat'], long=resp['long'], requested_datetime=resp[
                                                 "requested_datetime"], status=resp['status'],
                                             media_url=resp["media_url"], agency_responsible=resp["agency_responsible"])
                    solicitud.save()
            except Exception as e:
                print(str(e))

    context['response'] = response

    return JsonResponse(context)


@login_required(login_url='index')
def lista_solicitud(request):
    form = SolicitudFilterForm(request.GET)
    solicitudes = ListaOpen311.objects.all()

    if form.is_valid():
        descripcion = form.cleaned_data.get('descripcion')
        imagen = form.cleaned_data.get('imagen')
        direccion = form.cleaned_data.get('direccion')
        fecha = form.cleaned_data.get('fecha')

        if descripcion:
            solicitudes = solicitudes.filter(
                descripcion__icontains=descripcion)
        if imagen is not None and imagen != "":
            if imagen == "True":
                solicitudes = solicitudes.filter(media_url__isnull=False)
            else:
                solicitudes = solicitudes.filter(media_url__isnull=True)
        if direccion:
            solicitudes = solicitudes.filter(address__icontains=direccion)
        if fecha:
            fecha_parsed = parse_date(str(fecha))
            if fecha_parsed:
                solicitudes = solicitudes.filter(
                    requested_datetime__date=fecha_parsed)
        if not solicitudes.exists():
            messages.info(
                request, 'No se encontraron resultados con los filtros seleccionados.')

    paginator = Paginator(solicitudes, 15)

    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Obtener parámetros GET excepto el 'page' para mantener los filtros
    filters = request.GET.copy()
    if 'page' in filters:
        del filters['page']

    return render(request, 'open311/lista_solicitudes.html', {'solicitudes': page_obj, 'form': form, 'filters': filters})


def generar_reporte_excel(request):

    solicitudes = ListaOpen311.objects.all()

    wb = Workbook()
    ws = wb.active

    header_style = Alignment(
        horizontal='center', vertical='center', wrap_text=True)
    header_font = Font(bold=True)

    headers = ["Request ID", "Descripción", "Dirección", "Latitud", "Longitud",
               "Fecha de Creación", "Estado", "URL de Medios", "Agencia Responsable", "Activo"]
    for col_num, header_text in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        header_cell = ws[f"{col_letter}1"]
        header_cell.value = header_text
        header_cell.alignment = header_style
        header_cell.font = header_font

    cell_style = Alignment(horizontal='center',
                           vertical='center', wrap_text=True)

    for solicitud in solicitudes:

        requested_datetime = solicitud.requested_datetime.replace(tzinfo=None)

        row = [
            solicitud.request_id, solicitud.descripcion, solicitud.address,
            solicitud.lat, solicitud.long, requested_datetime, solicitud.status,
            solicitud.media_url, solicitud.agency_responsible, solicitud.activo
        ]

        ws.append(row)
        for col_num in range(1, len(row) + 1):
            col_letter = get_column_letter(col_num)
            cell = ws[f"{col_letter}{ws.max_row}"]
            cell.alignment = cell_style

    for col_num, _ in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].auto_size = True

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reporte_solicitudes.xlsx'

    wb.save(response)

    return response
